import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from pathlib import Path

class AttendanceSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("نظام تسجيل حضور وغياب الموظفين")
        self.root.geometry("1000x700")
        self.root.configure(bg='#2c3e50')

        # ملفات Excel
        self.attendance_file = "attendance_records.xlsx"
        self.employees_file = "employees_data.xlsx"

        # إنشاء الملفات إذا لم تكن موجودة
        self.initialize_excel_files()

        # الواجهة الرئيسية
        self.create_main_interface()

    def initialize_excel_files(self):
        """إنشاء ملفات Excel إذا لم تكن موجودة"""
        # ملف الموظفين
        if not os.path.exists(self.employees_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "الموظفين"
            ws.append(["رمز الموظف", "اسم الموظف", "القسم", "تاريخ التسجيل"])
            # تنسيق العناوين
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            wb.save(self.employees_file)

        # ملف الحضور
        if not os.path.exists(self.attendance_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "سجل الحضور"
            ws.append(["التاريخ", "الوقت", "رمز الموظف", "اسم الموظف", "نوع العملية", "ملاحظات"])
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            wb.save(self.attendance_file)

    def create_main_interface(self):
        """إنشاء الواجهة الرئيسية"""
        # العنوان
        title_frame = tk.Frame(self.root, bg='#34495e', height=80)
        title_frame.pack(fill='x', pady=10, padx=10)

        title_label = tk.Label(title_frame, text="🏢 نظام تسجيل حضور وغياب الموظفين",
                               font=('Arial', 24, 'bold'), bg='#34495e', fg='white')
        title_label.pack(pady=20)

        # الإطار الرئيسي
        main_frame = tk.Frame(self.root, bg='#2c3e50')
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # تقسيم الواجهة
        left_frame = tk.Frame(main_frame, bg='#34495e', width=400)
        left_frame.pack(side='left', fill='both', expand=True, padx=5)

        right_frame = tk.Frame(main_frame, bg='#34495e', width=580)
        right_frame.pack(side='right', fill='both', expand=True, padx=5)

        # القسم الأيسر - تسجيل الحضور
        self.create_attendance_section(left_frame)

        # القسم الأيمن - التقارير والإحصائيات
        self.create_reports_section(right_frame)

    def create_attendance_section(self, parent):
        """قسم تسجيل الحضور والغياب"""
        # إضافة موظف جديد
        add_emp_frame = tk.LabelFrame(parent, text="إضافة موظف جديد", 
                                      font=('Arial', 12, 'bold'), bg='#34495e', fg='white', pady=10)
        add_emp_frame.pack(fill='x', padx=10, pady=10)

        tk.Label(add_emp_frame, text="رمز الموظف:", bg='#34495e', fg='white').grid(row=0, column=1, padx=5, pady=5)
        self.emp_code_entry = tk.Entry(add_emp_frame, font=('Arial', 11), width=20)
        self.emp_code_entry.grid(row=0, column=0, padx=5, pady=5)

        tk.Label(add_emp_frame, text="اسم الموظف:", bg='#34495e', fg='white').grid(row=1, column=1, padx=5, pady=5)
        self.emp_name_entry = tk.Entry(add_emp_frame, font=('Arial', 11), width=20)
        self.emp_name_entry.grid(row=1, column=0, padx=5, pady=5)

        tk.Label(add_emp_frame, text="القسم:", bg='#34495e', fg='white').grid(row=2, column=1, padx=5, pady=5)
        self.emp_dept_entry = tk.Entry(add_emp_frame, font=('Arial', 11), width=20)
        self.emp_dept_entry.grid(row=2, column=0, padx=5, pady=5)

        tk.Button(add_emp_frame, text="➕ إضافة موظف", command=self.add_employee,
                 bg='#27ae60', fg='white', font=('Arial', 11, 'bold'), width=20).grid(row=3, column=0, columnspan=2, pady=10)

        # تسجيل الحضور/الانصراف
        attendance_frame = tk.LabelFrame(parent, text="تسجيل الحضور والانصراف",
                                        font=('Arial', 12, 'bold'), bg='#34495e', fg='white', pady=10)
        attendance_frame.pack(fill='x', padx=10, pady=10)

        tk.Label(attendance_frame, text="رمز الموظف:", bg='#34495e', fg='white').grid(row=0, column=1, padx=5, pady=5)
        self.check_code_entry = tk.Entry(attendance_frame, font=('Arial', 11), width=20)
        self.check_code_entry.grid(row=0, column=0, padx=5, pady=5)

        tk.Button(attendance_frame, text="✅ تسجيل حضور", command=lambda: self.record_attendance("حضور"),
                 bg='#3498db', fg='white', font=('Arial', 11, 'bold'), width=20).grid(row=1, column=1, pady=5)

        tk.Button(attendance_frame, text="🚪 تسجيل انصراف", command=lambda: self.record_attendance("انصراف"),
                 bg='#e74c3c', fg='white', font=('Arial', 11, 'bold'), width=20).grid(row=1, column=0, pady=5)

        # تسجيل إجازة/إذن
        leave_frame = tk.LabelFrame(parent, text="تسجيل إجازة أو إذن",
                                   font=('Arial', 12, 'bold'), bg='#34495e', fg='white', pady=10)
        leave_frame.pack(fill='x', padx=10, pady=10)

        tk.Label(leave_frame, text="رمز الموظف:", bg='#34495e', fg='white').grid(row=0, column=1, padx=5, pady=5)
        self.leave_code_entry = tk.Entry(leave_frame, font=('Arial', 11), width=20)
        self.leave_code_entry.grid(row=0, column=0, padx=5, pady=5)

        tk.Label(leave_frame, text="نوع الطلب:", bg='#34495e', fg='white').grid(row=1, column=1, padx=5, pady=5)
        self.leave_type_var = tk.StringVar(value="إجازة مرضية")
        leave_menu = ttk.Combobox(leave_frame, textvariable=self.leave_type_var,
                                 values=["إجازة مرضية", "إجازة عارضة", "إذن خروج", "إجازة رسمية"],
                                 font=('Arial', 10), width=18, state='readonly')
        leave_menu.grid(row=1, column=0, padx=5, pady=5)

        tk.Label(leave_frame, text="ملاحظات:", bg='#34495e', fg='white').grid(row=2, column=1, padx=5, pady=5)
        self.leave_notes_entry = tk.Entry(leave_frame, font=('Arial', 11), width=20)
        self.leave_notes_entry.grid(row=2, column=0, padx=5, pady=5)

        tk.Button(leave_frame, text="📝 تسجيل الطلب", command=self.record_leave,
                 bg='#f39c12', fg='white', font=('Arial', 11, 'bold'), width=20).grid(row=3, column=0, columnspan=2, pady=10)

    def create_reports_section(self, parent):
        """قسم التقارير والإحصائيات"""
        # التقرير الشهري
        report_frame = tk.LabelFrame(parent, text="التقرير الشهري",
                                    font=('Arial', 12, 'bold'), bg='#34495e', fg='white', pady=10)
        report_frame.pack(fill='x', padx=10, pady=10)

        tk.Label(report_frame, text="اختر الشهر:", bg='#34495e', fg='white').grid(row=0, column=2, padx=5, pady=5)
        self.month_var = tk.StringVar(value=str(datetime.now().month))
        month_menu = ttk.Combobox(report_frame, textvariable=self.month_var,
                                 values=[str(i) for i in range(1, 13)],
                                 font=('Arial', 10), width=10, state='readonly')
        month_menu.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(report_frame, text="السنة:", bg='#34495e', fg='white').grid(row=0, column=0, padx=5, pady=5)
        self.year_var = tk.StringVar(value=str(datetime.now().year))
        year_entry = tk.Entry(report_frame, textvariable=self.year_var, font=('Arial', 11), width=10)
        year_entry.grid(row=0, column=-1, padx=5, pady=5)

        tk.Button(report_frame, text="📊 إنشاء تقرير شهري Excel", command=self.generate_monthly_report,
                 bg='#16a085', fg='white', font=('Arial', 11, 'bold'), width=25).grid(row=1, column=0, columnspan=3, pady=10)

        # بحث عن موظف
        search_frame = tk.LabelFrame(parent, text="البحث عن موظف",
                                    font=('Arial', 12, 'bold'), bg='#34495e', fg='white', pady=10)
        search_frame.pack(fill='x', padx=10, pady=10)

        tk.Label(search_frame, text="رمز الموظف:", bg='#34495e', fg='white').grid(row=0, column=1, padx=5, pady=5)
        self.search_code_entry = tk.Entry(search_frame, font=('Arial', 11), width=15)
        self.search_code_entry.grid(row=0, column=0, padx=5, pady=5)

        tk.Button(search_frame, text="🔍 بحث", command=self.search_employee,
                 bg='#8e44ad', fg='white', font=('Arial', 11, 'bold'), width=15).grid(row=0, column=-1, pady=5, padx=5)

        # منطقة عرض النتائج
        results_frame = tk.LabelFrame(parent, text="النتائج والسجلات",
                                     font=('Arial', 12, 'bold'), bg='#34495e', fg='white')
        results_frame.pack(fill='both', expand=True, padx=10, pady=10)

        self.results_text = scrolledtext.ScrolledText(results_frame, width=60, height=15,
                                                     font=('Arial', 10), bg='#ecf0f1', wrap=tk.WORD)
        self.results_text.pack(padx=10, pady=10, fill='both', expand=True)

        # زر عرض جميع السجلات
        tk.Button(parent, text="📋 عرض آخر 20 سجل", command=self.show_recent_records,
                 bg='#2c3e50', fg='white', font=('Arial', 11, 'bold'), width=25).pack(pady=5)

    def add_employee(self):
        """إضافة موظف جديد"""
        code = self.emp_code_entry.get().strip()
        name = self.emp_name_entry.get().strip()
        dept = self.emp_dept_entry.get().strip()

        if not code or not name:
            messagebox.showerror("خطأ", "يرجى إدخال رمز واسم الموظف")
            return

        wb = openpyxl.load_workbook(self.employees_file)
        ws = wb.active

        # التحقق من عدم تكرار الرمز
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == code:
                messagebox.showerror("خطأ", "رمز الموظف موجود بالفعل!")
                wb.close()
                return

        ws.append([code, name, dept, datetime.now().strftime("%Y-%m-%d")])
        wb.save(self.employees_file)
        wb.close()

        messagebox.showinfo("نجح", f"تم إضافة الموظف {name} بنجاح!")
        self.emp_code_entry.delete(0, tk.END)
        self.emp_name_entry.delete(0, tk.END)
        self.emp_dept_entry.delete(0, tk.END)

    def get_employee_name(self, code):
        """الحصول على اسم الموظف من الرمز"""
        wb = openpyxl.load_workbook(self.employees_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == code:
                wb.close()
                return row[1]
        wb.close()
        return None

    def record_attendance(self, record_type):
        """تسجيل الحضور أو الانصراف"""
        code = self.check_code_entry.get().strip()

        if not code:
            messagebox.showerror("خطأ", "يرجى إدخال رمز الموظف")
            return

        name = self.get_employee_name(code)
        if not name:
            messagebox.showerror("خطأ", "الموظف غير موجود! يرجى إضافته أولاً")
            return

        wb = openpyxl.load_workbook(self.attendance_file)
        ws = wb.active

        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")

        ws.append([current_date, current_time, code, name, record_type, ""])
        wb.save(self.attendance_file)
        wb.close()

        icon = "✅" if record_type == "حضور" else "🚪"
        messagebox.showinfo("نجح", f"{icon} تم تسجيل {record_type} للموظف {name}\nالوقت: {current_time}")
        self.check_code_entry.delete(0, tk.END)

    def record_leave(self):
        """تسجيل إجازة أو إذن"""
        code = self.leave_code_entry.get().strip()
        leave_type = self.leave_type_var.get()
        notes = self.leave_notes_entry.get().strip()

        if not code:
            messagebox.showerror("خطأ", "يرجى إدخال رمز الموظف")
            return

        name = self.get_employee_name(code)
        if not name:
            messagebox.showerror("خطأ", "الموظف غير موجود!")
            return

        wb = openpyxl.load_workbook(self.attendance_file)
        ws = wb.active

        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")

        ws.append([current_date, current_time, code, name, leave_type, notes])
        wb.save(self.attendance_file)
        wb.close()

        messagebox.showinfo("نجح", f"تم تسجيل {leave_type} للموظف {name}")
        self.leave_code_entry.delete(0, tk.END)
        self.leave_notes_entry.delete(0, tk.END)

    def search_employee(self):
        """البحث عن سجلات موظف معين"""
        code = self.search_code_entry.get().strip()

        if not code:
            messagebox.showerror("خطأ", "يرجى إدخال رمز الموظف")
            return

        name = self.get_employee_name(code)
        if not name:
            messagebox.showerror("خطأ", "الموظف غير موجود!")
            return

        wb = openpyxl.load_workbook(self.attendance_file)
        ws = wb.active

        records = []
        days_present = 0
        days_absent = 0
        leaves_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == code:
                records.append(row)
                if row[4] == "حضور":
                    days_present += 1
                elif "إجازة" in row[4] or "إذن" in row[4]:
                    leaves_count += 1

        wb.close()

        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, f"{'='*60}\n")
        self.results_text.insert(tk.END, f"   سجل الموظف: {name} (رمز: {code})\n")
        self.results_text.insert(tk.END, f"{'='*60}\n\n")
        self.results_text.insert(tk.END, f"📊 الإحصائيات:\n")
        self.results_text.insert(tk.END, f"   • أيام الحضور: {days_present}\n")
        self.results_text.insert(tk.END, f"   • الإجازات والأذونات: {leaves_count}\n")
        self.results_text.insert(tk.END, f"   • إجمالي السجلات: {len(records)}\n\n")
        self.results_text.insert(tk.END, f"{'='*60}\n")
        self.results_text.insert(tk.END, f"آخر 10 سجلات:\n")
        self.results_text.insert(tk.END, f"{'='*60}\n\n")

        for record in records[-10:]:
            self.results_text.insert(tk.END, f"📅 التاريخ: {record[0]}  |  ⏰ الوقت: {record[1]}\n")
            self.results_text.insert(tk.END, f"   النوع: {record[4]}\n")
            if record[5]:
                self.results_text.insert(tk.END, f"   ملاحظات: {record[5]}\n")
            self.results_text.insert(tk.END, f"{'-'*60}\n")

    def show_recent_records(self):
        """عرض آخر السجلات"""
        wb = openpyxl.load_workbook(self.attendance_file)
        ws = wb.active

        records = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, f"{'='*60}\n")
        self.results_text.insert(tk.END, f"   آخر 20 سجل حضور وانصراف\n")
        self.results_text.insert(tk.END, f"{'='*60}\n\n")

        for record in records[-20:]:
            self.results_text.insert(tk.END, f"📅 {record[0]} | ⏰ {record[1]}\n")
            self.results_text.insert(tk.END, f"   الموظف: {record[3]} ({record[2]})\n")
            self.results_text.insert(tk.END, f"   العملية: {record[4]}\n")
            if record[5]:
                self.results_text.insert(tk.END, f"   ملاحظات: {record[5]}\n")
            self.results_text.insert(tk.END, f"{'-'*60}\n")

    def generate_monthly_report(self):
        """إنشاء تقرير شهري Excel"""
        try:
            month = int(self.month_var.get())
            year = int(self.year_var.get())
        except:
            messagebox.showerror("خطأ", "يرجى إدخال شهر وسنة صحيحين")
            return

        # قراءة بيانات الحضور
        wb_att = openpyxl.load_workbook(self.attendance_file)
        ws_att = wb_att.active

        # قراءة بيانات الموظفين
        wb_emp = openpyxl.load_workbook(self.employees_file)
        ws_emp = wb_emp.active

        # جمع بيانات الموظفين
        employees = {}
        for row in ws_emp.iter_rows(min_row=2, values_only=True):
            employees[row[0]] = {"name": row[1], "dept": row[2], "attendance": {}, "leaves": 0}

        # معالجة سجلات الحضور
        for row in ws_att.iter_rows(min_row=2, values_only=True):
            date_str = str(row[0])
            try:
                record_date = datetime.strptime(date_str, "%Y-%m-%d")
                if record_date.month == month and record_date.year == year:
                    emp_code = row[2]
                    if emp_code in employees:
                        date_key = record_date.strftime("%Y-%m-%d")
                        if date_key not in employees[emp_code]["attendance"]:
                            employees[emp_code]["attendance"][date_key] = []
                        employees[emp_code]["attendance"][date_key].append({
                            "time": row[1],
                            "type": row[4],
                            "notes": row[5]
                        })
                        if "إجازة" in row[4] or "إذن" in row[4]:
                            employees[emp_code]["leaves"] += 1
            except:
                continue

        wb_att.close()
        wb_emp.close()

        # إنشاء التقرير
        wb_report = openpyxl.Workbook()
        ws_report = wb_report.active
        ws_report.title = f"تقرير {month}-{year}"

        # العناوين
        headers = ["رمز الموظف", "اسم الموظف", "القسم", "أيام الحضور", "الإجازات والأذونات", "إجمالي السجلات"]
        ws_report.append(headers)

        # تنسيق العناوين
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        for cell in ws_report[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # إضافة البيانات
        for emp_code, data in employees.items():
            days_present = len([d for d, records in data["attendance"].items() 
                               if any(r["type"] == "حضور" for r in records)])
            total_records = sum(len(records) for records in data["attendance"].values())

            ws_report.append([
                emp_code,
                data["name"],
                data["dept"],
                days_present,
                data["leaves"],
                total_records
            ])

        # تنسيق الأعمدة
        ws_report.column_dimensions['A'].width = 15
        ws_report.column_dimensions['B'].width = 25
        ws_report.column_dimensions['C'].width = 20
        ws_report.column_dimensions['D'].width = 15
        ws_report.column_dimensions['E'].width = 20
        ws_report.column_dimensions['F'].width = 18

        # حفظ التقرير
        report_filename = f"تقرير_شهر_{month}_{year}.xlsx"
        wb_report.save(report_filename)
        wb_report.close()

        messagebox.showinfo("نجح", f"تم إنشاء التقرير الشهري بنجاح!\nاسم الملف: {report_filename}")

        # عرض ملخص في منطقة النتائج
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, f"{'='*60}\n")
        self.results_text.insert(tk.END, f"   التقرير الشهري - {month}/{year}\n")
        self.results_text.insert(tk.END, f"{'='*60}\n\n")
        self.results_text.insert(tk.END, f"✅ تم إنشاء التقرير بنجاح\n")
        self.results_text.insert(tk.END, f"📁 اسم الملف: {report_filename}\n")
        self.results_text.insert(tk.END, f"👥 عدد الموظفين: {len(employees)}\n")

# تشغيل البرنامج
if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceSystem(root)
    root.mainloop()
